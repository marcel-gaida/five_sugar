#!/usr/bin/env python
import subprocess
try:
    import openpyxl
except ImportError:
    # If not, install pandas using subprocess
    subprocess.run(["pip", "install", "openpyxl"])
    # Reimport
    import openpyxl

# Open the Excel spreadsheet
wb = openpyxl.load_workbook('merged.xlsx')

# Get the active sheet
sheet = wb.active

# Iterate through the rows in the sheet, starting from row 9
for row in sheet.iter_rows(min_row=9):
    # Iterate through the columns in the row
    for cell in row:
        # Get the cell value
        cell_value = cell.value

        # Delete "M1=" and " (fraction of A overlapping B)" from the cell, and keep only the number
        if cell_value is not None:
            m1_index = cell_value.find("M1=")
            if m1_index != -1:
                cell_value = cell_value[m1_index + 3:]
                cell_value = cell_value.replace(" (fraction of A overlapping B)", "")
                cell.value = cell_value

for row in sheet.iter_rows(min_row=10):
    # Iterate through the columns in the row
    for cell in row:
        # Get the cell value
        cell_value = cell.value

        # Delete "M2=" and " (fraction of B overlapping A)" from the cell, and keep only the number
        if cell_value is not None:
            m1_index = cell_value.find("M2=")
            if m1_index != -1:
                cell_value = cell_value[m1_index + 3:]
                cell_value = cell_value.replace(" (fraction of B overlapping A)", "")
                cell.value = cell_value

# Save the changes to the spreadsheet
wb.save('merged.xlsx')

